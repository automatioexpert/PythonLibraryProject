import openpyxl

wb = openpyxl.load_workbook("TestData.xlsx")
sh1 = wb['Sheet1']

rows = sh1.max_row
column = sh1.max_column

for i in range(1, rows + 1):
    for j in range(1, column + 1):
        print(sh1.cell(i, j).value)

# Write data to excel

sh1.cell(row=5, column=1, value='Pytest')
sh1.cell(row=12, column=1, value='JUnit')
sh1.cell(row=14, column=1, value='TestNG')
sh1.cell(row=15, column=1, value='NUnit')
wb.save("Report.xlsx")

"""OUTPUT
username
password
stand_user
secret_sauce
stand_user1
secret_sauce1
stand_user2
secret_sauce2
stand_user3
secret_sauce3
stand_user4
secret_sauce4
stand_user5
secret_sauce5
stand_user6
secret_sauce6
stand_user7
secret_sauce7
stand_user8
secret_sauce8
"""
