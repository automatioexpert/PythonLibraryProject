import openpyxl

wb=openpyxl.load_workbook("TestData.xlsx")
print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'>
sheets=wb.sheetnames
print(sheets)
print(wb.active.title) #Current active sheet
sh1=wb['Sheet1']
print(type(sh1))
data=sh1['A2'].value
print(data)
print(wb['Sheet1']['A4'].value)

#option2:
print(sh1.cell(1,2).value)

#option3 : print all data
